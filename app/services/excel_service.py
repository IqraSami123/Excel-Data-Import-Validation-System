from openpyxl import load_workbook


REQUIRED_COLUMNS = {
    "name",
    "email",
    "phone",
    "age",
    "city",
    "status",
}


def read_excel_rows(file_path: str):
    ### this function will read the excel file and yield the rows one by one as a dictionary with the header as the key and the cell value as the value###
    workbook = load_workbook(
        filename=file_path,
        read_only=True,     #it loads the workbook just in readable form
        data_only=True,      #it loads the workbook with only the data and not the formulas
    )

    worksheet = workbook.active   # to get the active sheet of the workbook
    header_row = next(worksheet.iter_rows(values_only=True))    # this will get the first row of the sheet which is the header row

    headers = [
        str(header).strip().lower()  # this will convert the header to string and then strip the whitespace and then convert it to lowercase to validate the headers
        if header is not None
        else ""
        for header in header_row
    ]

    missing_columns = REQUIRED_COLUMNS - set(headers)

    if missing_columns:
        workbook.close()

        raise ValueError(
            f"Missing required columns: {', '.join(sorted(missing_columns))}"
        )

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        yield row_data        #it will generate the row ony by one not in once

    workbook.close()